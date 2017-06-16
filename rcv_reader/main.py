import PyPDF2

# pdfFileObj = open('cameronmoll_bio.pdf', 'rb')
# pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
#
# print(pdfReader.numPages)
# pageObj = pdfReader.getPage(0)
# print(pageObj.extractText())

import ffile
import os, os.path
import openpyxl

def read_pdf(filename):
    pass

def make_pdf_file(pdfwriter, pdf_name):
    pdf_filename = pdf_name + '.pdf'
    if os.path.isfile(pdf_filename):
        new_filename = ''
        for i in range(100):

            new_filename = pdf_name + '_' + str(i) + '.pdf'
            if os.path.isfile(new_filename):
                continue
            else:
                pdf_filename = new_filename
                break

    pdfOutputFile = open(pdf_filename, 'wb')
    pdfwriter.write(pdfOutputFile)
    pdfOutputFile.close()

def get_pdf_index_dic():
    index_dic = {}

    ffile.move_dir("unread_pdf")

    try:
        wb = openpyxl.load_workbook('index.xlsx')
    except FileNotFoundError:
        print("index.xlsx was not found")
        return 0

    sheet_list = wb.get_sheet_names()
    sheetname = sheet_list[0]
    sheet = wb.get_sheet_by_name(sheetname)

    max_columns = sheet.max_column

    cur_year = ''
    cur_month = ''

    repeat_page_symbol = '-'


    for col in range(1, max_columns + 1, 4):
        pdf_name = str(sheet.cell(row=1, column=col).value)
        index_dic[pdf_name] = []

        num_pages = sheet.cell(row=2, column=col).value
        if num_pages == None:
            msg = "No num of pages " + pdf_name + " in column " + str(col)
            raise Exception(msg)

        pages_col_num = col + 1
        year_col = col + 1
        month_col = col + 2
        day_rcv_col = col + 3

        year = ''
        month = ''
        day_rcv = ''

        for page in range(1, num_pages + 1):
            year_cell = sheet.cell(row = page, column=year_col).value

            if year_cell != None:
                if year_cell < 10:
                    year = "0" + str(year_cell)
                else:
                    year = str(year_cell)

            month_cell = sheet.cell(row = page, column=month_col).value

            if month_cell != None:
                if month_cell < 10:
                    month = "0" + str(month_cell)
                else:
                    month = str(month_cell)

            day_rcv_cell = sheet.cell(row = page, column=day_rcv_col).value

            day_rcv = str(day_rcv_cell)

            if day_rcv == repeat_page_symbol:
                rcv_name = day_rcv
            else:
                rcv_name = year + month + day_rcv
            index_dic[pdf_name].append(rcv_name)

    ffile.dir_back()

    return index_dic

def process_pdfs():
    index_dic = get_pdf_index_dic()
    print("INDEX", index_dic)

    for pdf_name, rcv_name_list in index_dic.items():

        ffile.move_dir('unread_pdf')

        msg = "Processing " + pdf_name
        print(msg)

        
        pdf_filename = pdf_name + '.pdf'

        print(pdf_filename)

        try:
            pdfFile = open(pdf_filename, 'rb')
        except FileNotFoundError:
            print("FILE NOT FOUND " + pdf_filename)
            print(os.getcwd())
            continue

        pdfReader = PyPDF2.PdfFileReader(pdfFile)

        ffile.dir_back()


        ffile.move_dir('proc_pdf/')

        pdf_pages = pdfReader.numPages

        num_pdf_names = len(rcv_name_list)
        if pdf_pages != num_pdf_names:
            raise Exception('num_pages != pdf_pages of file ' + pdf_filename)

        prev_rcv_filename = ""
        prev_pdfwriter = PyPDF2.PdfFileWriter()


        for i, rcv_name in enumerate(rcv_name_list):
            page = i + 1
            if rcv_name != "-":
                if page != 1:
                    # print(page, "TEST",prev_pdfwriter, prev_rcv_filename)
                    #Write previous RCV PDF
                    make_pdf_file(prev_pdfwriter, prev_rcv_filename)

                prev_rcv_filename = 'RCV' + rcv_name
                prev_pdfwriter = PyPDF2.PdfFileWriter()


            pageObj = pdfReader.getPage(page - 1)
            prev_pdfwriter.addPage(pageObj)
        
        make_pdf_file(prev_pdfwriter, prev_rcv_filename)

        ffile.dir_back()

process_pdfs()