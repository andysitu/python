import ffile
import openpyxl, os
from openpyxl.utils import column_index_from_string

import excel_maker

from os.path import isfile, join
import operator

def open_data_excel():
    ffile.move_dir("excel_data")

    mypath = '.'
    cust_id_letter = 'G'
    cbm_col_letter = 'O'
    container_id_letter = 'P'


    files_list = [f for f in os.listdir(mypath) if isfile(join(mypath, f))]

    total_data_dic = {}

    for filename in files_list:
        filename_encoded = filename.encode('utf8')
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.get_sheet_names()

        file_data_dic = total_data_dic[filename_encoded] = {}

        for sheetname in sheetname_list:
            sheet = wb.get_sheet_by_name(sheetname)
            sheetname_encoded = sheetname.encode('utf8')
            data_dic = {}
            file_data_dic[sheetname_encoded] = data_dic
            max_row = sheet.max_row

            prev_pallet_id = ""
            for row in range(2, max_row + 1):
                row_str = str(row)

                cust_id = sheet[cust_id_letter + row_str].value
                if cust_id == None:
                    continue

                if cust_id not in data_dic:
                    data_dic[cust_id] = []

                try:
                    cbm = float(sheet[cbm_col_letter + row_str].value)
                except (ValueError, TypeError):
                    cbm = None
                pallet_id = sheet[pallet_id_letter + row_str].value

                if pallet_id == None:
                    pallet_id = prev_pallet_id

                data_dic[cust_id].append({"cbm": cbm, "pallet_id": pallet_id})

                prev_pallet_id = pallet_id
    ffile.dir_back()

    return total_data_dic


def make_summarize_data_excel(total_data_dic):

    for excel_filename_encoded, file_data_dic in total_data_dic.items():
        excel_dic = {}
        excel_filename = excel_filename_encoded.decode('utf8')
        wb = openpyxl.Workbook()

        for sheetname_encoded, data_dic in file_data_dic.items():

            pallet_count_dic = {}
            cust_cbm_dic = {}
            total_cbm = 0.00

            sheetname = sheetname_encoded.decode('utf8')
            excel_dic[sheetname] = []

            excel_dic[sheetname] = []
            for cust_id, data_list in data_dic.items():
                cust_cbm_dic[cust_id] = 0
                for d in data_list:
                    cbm = d["cbm"]
                    pallet_id = d["pallet_id"]

                    if cbm != None:
                        total_cbm += cbm
                        cust_cbm_dic[cust_id] += cbm
                    if pallet_id != None:
                        pallet_count_dic[pallet_id] = True

            sheet_list = []
            excel_dic[sheetname] = sheet_list

            # List count of pallet_ids
            pallet_count_list = ["# Container", len(pallet_count_dic)]
            sheet_list.append(pallet_count_list)
            sheet_list.append([])

            # total cbm
            cbm_count_list = ["Total CBM", total_cbm]
            sheet_list.append(cbm_count_list)
            sheet_list.append([])

            # list by cust_id cbm
            sheet_list.append(["Customer Sort by CBM"])
            cust_cbm_dic
            sorted_cust_cbm = sorted(cust_cbm_dic.items(), key=operator.itemgetter(1), reverse=True)
            for cust_cbm_tup in sorted_cust_cbm:
                cust_id = cust_cbm_tup[0]
                cbm = cust_cbm_tup[1]
                cust_cbm_list = [cust_id, cbm]
                sheet_list.append(cust_cbm_list)


        # -5 to remove ".xlsx" from original filename
        new_filename = excel_filename[:-5] + str(" summary.xlsx")

        excel_maker.make_excel_file(excel_dic, new_filename, "excel_files")

total_data_dic = open_data_excel()
make_summarize_data_excel(total_data_dic)